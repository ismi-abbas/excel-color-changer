import os
from openpyxl import load_workbook
from openpyxl.drawing.image import Image


# This one only to replace image, cause ExcelJS doesnt s
def process_xlsx_file(file_path):
    try:
        workbook = load_workbook(file_path)

        # Loop through each sheet in the workbook
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]

            # Check if the sheet has images
            if worksheet._images:
                for image_id, image in enumerate(worksheet._images):
                    worksheet._images.remove(image)

                # Specify the path to the new image file
                new_image_file_path = os.path.join(
                    os.path.dirname(__file__), "POTF.png"
                )

                # Add the new image to the sheet
                potf_image = Image(new_image_file_path)
                potf_image.height = 75
                potf_image.width = 70

                worksheet.add_image(potf_image, "B1")

        # Save the modified workbook
        workbook.save(file_path)
        print(f"Processed and saved {file_path}")
    except FileNotFoundError:
        print(f"Error: File not found - {file_path}")
    except Exception as e:
        print(f"Error processing {file_path}: {str(e)}")


# Get a list of all XLSX files in the current directory
current_dir = os.path.join(os.path.dirname(__file__), "files")
for file in os.listdir(current_dir):
    file_path = os.path.join(current_dir, file)

    # Check if the file is an XLSX file
    if os.path.isfile(file_path) and file_path.lower().endswith(".xlsx"):
        try:
            process_xlsx_file(file_path)
            print(f"Processed and saved {file_path}")
        except Exception as e:
            print(f"Error processing {file_path}: {str(e)}")
